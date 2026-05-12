"""
PhantomFeed — Export API Routes

GET /api/v1/export/items.csv              Threat items as CSV
GET /api/v1/export/items.json             Threat items as JSON
GET /api/v1/export/iocs.txt               IOC list plain text
GET /api/v1/export/iocs.csv               IOC list CSV
GET /api/v1/export/iocs.stix              IOC list as STIX 2.1 bundle
GET /api/v1/clients/{id}/export/remediation.csv   Remediation tracker CSV
GET /api/v1/clients/{id}/export/remediation.xlsx  Remediation tracker XLSX
GET /api/v1/clients/{id}/export/detection-rules.zip  SPL/KQL/Sigma ZIP
POST /api/v1/clients/{id}/export/push-rules-github   Push rules to GitHub
"""

import csv
import io
import json
import re
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import JSONResponse, StreamingResponse

from auth.auth import decode_token

router = APIRouter()


async def _require_auth(
    request: Request,
    token: Optional[str] = Query(None),
) -> dict:
    """Accept JWT from ?token= query param or Authorization: Bearer header."""
    raw = token
    if not raw:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            raw = auth_header[7:]
    if not raw:
        raise HTTPException(401, "Not authenticated")
    payload = decode_token(raw)
    user_id = payload.get("sub")
    if not user_id:
        raise HTTPException(401, "Invalid token")
    from db import database as db
    user = await db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(401, "User not found")
    return user


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")[:60]


def _days_remaining(due_date: str) -> Optional[int]:
    if not due_date:
        return None
    try:
        due = datetime.strptime(due_date[:10], "%Y-%m-%d")
        return (due - datetime.now(timezone.utc).replace(tzinfo=None)).days
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Shared item query helper
# ---------------------------------------------------------------------------

async def _query_items(
    severity: Optional[str],
    category: Optional[str],
    feed_id: Optional[str],
    search: Optional[str],
    client_id: Optional[str],
    days: Optional[int],
    limit: int = 5000,
) -> list[dict]:
    from db import database as db
    items = await db.get_items(
        severity=severity,
        category=category,
        feed_id=feed_id,
        search=search,
        client_id=client_id,
        limit=limit,
    )
    if days:
        cutoff = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)).strftime("%Y-%m-%d")
        items = [i for i in items if (i.get("published_at") or "") >= cutoff]
    return items


# ---------------------------------------------------------------------------
# GET /export/items.csv
# ---------------------------------------------------------------------------

@router.get("/export/items.csv")
async def export_items_csv(
    severity: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    feed_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    client_id: Optional[str] = Query(None),
    days: Optional[int] = Query(None),
    user: dict = Depends(_require_auth),
):
    items = await _query_items(severity, category, feed_id, search, client_id, days)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "id", "title", "severity", "cvss", "risk_score", "vendor", "product",
        "feed_label", "category", "published_at", "url", "cve_ids", "tags",
        "compliance_cmmc", "compliance_nist",
    ])
    for item in items:
        tags = item.get("tags") or []
        if isinstance(tags, str):
            try:
                tags = json.loads(tags)
            except Exception:
                tags = []
        compliance = item.get("compliance_tags") or []
        if isinstance(compliance, str):
            try:
                compliance = json.loads(compliance)
            except Exception:
                compliance = []
        cves = item.get("cve_ids") or []
        if isinstance(cves, str):
            try:
                cves = json.loads(cves)
            except Exception:
                cves = []
        cmmc = [t for t in compliance if t.startswith("CMMC")]
        nist = [t for t in compliance if t.startswith("NIST")]
        writer.writerow([
            item.get("id", ""),
            item.get("title", ""),
            item.get("severity", ""),
            item.get("cvss", ""),
            item.get("risk_score", ""),
            item.get("vendor", ""),
            item.get("product", ""),
            item.get("feed_label", ""),
            item.get("category", ""),
            item.get("published_at", ""),
            item.get("url", ""),
            ";".join(cves),
            ";".join(str(t) for t in tags),
            ";".join(cmmc),
            ";".join(nist),
        ])

    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="phantomfeed_items.csv"'},
    )


# ---------------------------------------------------------------------------
# GET /export/items.json
# ---------------------------------------------------------------------------

@router.get("/export/items.json")
async def export_items_json(
    severity: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    feed_id: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    client_id: Optional[str] = Query(None),
    days: Optional[int] = Query(None),
    user: dict = Depends(_require_auth),
):
    items = await _query_items(severity, category, feed_id, search, client_id, days)
    content = json.dumps({"count": len(items), "items": items}, default=str, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="phantomfeed_items.json"'},
    )


# ---------------------------------------------------------------------------
# GET /export/iocs.txt
# ---------------------------------------------------------------------------

@router.get("/export/iocs.txt")
async def export_iocs_txt(
    days: int = Query(7),
    type: Optional[str] = Query(None, description="ip, hash, domain, url, or all"),
    user: dict = Depends(_require_auth),
):
    rows = await _get_ioc_rows(days, type)
    lines = [r["ioc_value"] for r in rows]
    content = "\n".join(lines).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/plain",
        headers={"Content-Disposition": 'attachment; filename="phantomfeed_iocs.txt"'},
    )


# ---------------------------------------------------------------------------
# GET /export/iocs.csv
# ---------------------------------------------------------------------------

@router.get("/export/iocs.csv")
async def export_iocs_csv(
    days: int = Query(7),
    type: Optional[str] = Query(None),
    user: dict = Depends(_require_auth),
):
    rows = await _get_ioc_rows(days, type)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["type", "value", "source", "threat_name", "confidence", "first_seen"])
    for r in rows:
        vt_name = r.get("vt_name", "") or ""
        greynoise = r.get("greynoise_classification", "") or ""
        abuseipdb = r.get("abuseipdb_score")
        confidence = "high" if (abuseipdb and abuseipdb > 50) or greynoise == "malicious" else "medium"
        writer.writerow([
            r.get("ioc_type", ""),
            r.get("ioc_value", ""),
            "PhantomFeed",
            vt_name or greynoise,
            confidence,
            (r.get("enriched_at") or "")[:10],
        ])
    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="phantomfeed_iocs.csv"'},
    )


# ---------------------------------------------------------------------------
# GET /export/iocs.stix
# ---------------------------------------------------------------------------

@router.get("/export/iocs.stix")
async def export_iocs_stix(
    days: int = Query(7),
    type: Optional[str] = Query(None),
    user: dict = Depends(_require_auth),
):
    rows = await _get_ioc_rows(days, type)
    now = datetime.now(timezone.utc).replace(tzinfo=None).strftime("%Y-%m-%dT%H:%M:%SZ")
    indicators = []
    for r in rows:
        ioc_type = r.get("ioc_type", "unknown")
        val = r.get("ioc_value", "")
        pattern = _stix_pattern(ioc_type, val)
        if not pattern:
            continue
        indicators.append({
            "type": "indicator",
            "spec_version": "2.1",
            "id": f"indicator--{uuid.uuid4()}",
            "created": now,
            "modified": now,
            "name": f"{ioc_type}: {val}",
            "pattern": pattern,
            "pattern_type": "stix",
            "valid_from": now,
            "indicator_types": ["malicious-activity"],
        })

    bundle = {
        "type": "bundle",
        "id": f"bundle--{uuid.uuid4()}",
        "objects": indicators,
    }
    content = json.dumps(bundle, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="phantomfeed_iocs.stix.json"'},
    )


def _stix_pattern(ioc_type: str, value: str) -> Optional[str]:
    if ioc_type == "ip":
        return f"[ipv4-addr:value = '{value}']"
    if ioc_type == "ipv6":
        return f"[ipv6-addr:value = '{value}']"
    if ioc_type == "domain":
        return f"[domain-name:value = '{value}']"
    if ioc_type == "url":
        return f"[url:value = '{value}']"
    if ioc_type in ("md5", "sha1", "sha256"):
        return f"[file:hashes.'{ioc_type.upper()}' = '{value}']"
    return None


async def _get_ioc_rows(days: int, ioc_type_filter: Optional[str]) -> list[dict]:
    from db import database as db
    cutoff = (datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)).isoformat()
    all_rows = await db.list_ioc_cache(limit=10000)
    rows = [r for r in all_rows if (r.get("enriched_at") or "") >= cutoff]
    if ioc_type_filter and ioc_type_filter != "all":
        type_set = set()
        if ioc_type_filter == "ip":
            type_set = {"ip", "ipv6"}
        elif ioc_type_filter == "hash":
            type_set = {"md5", "sha1", "sha256"}
        else:
            type_set = {ioc_type_filter}
        rows = [r for r in rows if r.get("ioc_type") in type_set]
    return rows


# ---------------------------------------------------------------------------
# GET /clients/{id}/export/remediation.csv
# ---------------------------------------------------------------------------

@router.get("/clients/{client_id}/export/remediation.csv")
async def export_remediation_csv(client_id: str, user: dict = Depends(_require_auth)):
    rows = await _get_remediation_rows(client_id)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "title", "severity", "cvss", "risk_score", "status",
        "assigned_to", "due_date", "days_remaining", "sla_breached", "cve_ids", "vendor",
    ])
    for r in rows:
        dr = _days_remaining(r.get("due_date"))
        sla_breached = "YES" if r.get("is_overdue") or (dr is not None and dr < 0) else "NO"
        writer.writerow([
            r.get("title", ""),
            r.get("severity", ""),
            r.get("cvss", ""),
            r.get("risk_score", ""),
            r.get("status", ""),
            r.get("assigned_to", ""),
            r.get("due_date", ""),
            dr if dr is not None else "",
            sla_breached,
            ";".join(r.get("cve_ids", [])),
            r.get("vendor", ""),
        ])
    content = buf.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="remediation_tracker.csv"'},
    )


# ---------------------------------------------------------------------------
# GET /clients/{id}/export/remediation.xlsx
# ---------------------------------------------------------------------------

@router.get("/clients/{client_id}/export/remediation.xlsx")
async def export_remediation_xlsx(client_id: str, user: dict = Depends(_require_auth)):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    rows = await _get_remediation_rows(client_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Remediation Tracker"

    headers = [
        "Title", "Severity", "CVSS", "Risk Score", "Status",
        "Assigned To", "Due Date", "Days Remaining", "SLA Breached", "CVE IDs", "Vendor",
    ]
    ws.append(headers)
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.font = Font(bold=True, color="FFFFFF")

    RED = PatternFill("solid", fgColor="FECACA")
    AMBER = PatternFill("solid", fgColor="FEF08A")
    GREEN = PatternFill("solid", fgColor="BBF7D0")

    for r in rows:
        dr = _days_remaining(r.get("due_date"))
        sla_breached = "YES" if r.get("is_overdue") or (dr is not None and dr < 0) else "NO"
        status = r.get("status", "")
        row_data = [
            r.get("title", "")[:200],
            r.get("severity", ""),
            r.get("cvss", ""),
            r.get("risk_score", ""),
            status,
            r.get("assigned_to", "") or "",
            r.get("due_date", ""),
            dr if dr is not None else "",
            sla_breached,
            ";".join(r.get("cve_ids", [])),
            r.get("vendor", "") or "",
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        # Color coding
        if status in ("patched",):
            fill = GREEN
        elif sla_breached == "YES" or (dr is not None and dr < 0):
            fill = RED
        elif dr is not None and dr <= 7:
            fill = AMBER
        else:
            fill = None
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    # Auto-size columns
    for col_idx, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="remediation_tracker.xlsx"'},
    )


async def _get_remediation_rows(client_id: str) -> list[dict]:
    """Join remediation items with their threat_item fields."""
    from db import database as db
    rems = await db.get_remediations(client_id)
    result = []
    for rem in rems:
        item_id = rem.get("item_id", "")
        threat_item = {}
        if item_id:
            ti = await db.get_item(item_id)
            if ti:
                threat_item = ti
        cve_ids = threat_item.get("cve_ids") or []
        if isinstance(cve_ids, str):
            try:
                cve_ids = json.loads(cve_ids)
            except Exception:
                cve_ids = []
        result.append({
            **rem,
            "title": threat_item.get("title", rem.get("item_id", "")),
            "severity": threat_item.get("severity", ""),
            "cvss": threat_item.get("cvss"),
            "risk_score": threat_item.get("risk_score"),
            "vendor": threat_item.get("vendor", ""),
            "cve_ids": cve_ids,
        })
    return result


# ---------------------------------------------------------------------------
# GET /clients/{id}/export/detection-rules.zip
# ---------------------------------------------------------------------------

SPL_TEMPLATE = """\
| search {search_terms}
| eval severity="{severity}", cve="{cves}", rule_name="{title}"
| table _time, src_ip, dest_ip, severity, cve, rule_name
"""

KQL_TEMPLATE = """\
// {title} — {severity}
// CVEs: {cves}
SecurityEvent
| where TimeGenerated > ago(24h)
| where {kql_filter}
| project TimeGenerated, Computer, EventID, Account, Activity
"""

SIGMA_TEMPLATE = """\
title: {title}
id: {rule_id}
status: experimental
description: {description}
references: {refs}
tags:
{tags}
logsource:
    product: windows
    service: sysmon
detection:
    selection:
        EventID: 1
        Image|contains: '{product}'
    condition: selection
level: {level}
"""


def _make_rules(item: dict) -> tuple[str, str, str]:
    """Generate SPL, KQL, Sigma for one item."""
    title = item.get("title", "Unknown")[:100]
    severity = item.get("severity", "MEDIUM")
    vendor = (item.get("vendor") or "").replace("'", "")
    product = (item.get("product") or vendor or "").replace("'", "")
    cves_list = item.get("cve_ids") or []
    if isinstance(cves_list, str):
        try:
            cves_list = json.loads(cves_list)
        except Exception:
            cves_list = []
    cves = ", ".join(cves_list) or "N/A"
    description = (item.get("description") or "")[:200].replace('"', "'")

    search_terms = f'"{vendor}"' if vendor else '"vulnerability"'
    kql_filter = f'Computer contains "{vendor}"' if vendor else 'EventID == 4625'
    level = "critical" if severity == "CRITICAL" else "high" if severity == "HIGH" else "medium"

    refs = f"    - {item.get('url', 'https://nvd.nist.gov')}" if item.get("url") else "    - https://nvd.nist.gov"
    tags_lines = "\n".join(f"    - {cve.lower().replace('-', '.')}" for cve in cves_list) or "    - attack.execution"
    rule_id = str(uuid.uuid4())

    spl = SPL_TEMPLATE.format(search_terms=search_terms, severity=severity, cves=cves, title=title)
    kql = KQL_TEMPLATE.format(title=title, severity=severity, cves=cves, kql_filter=kql_filter)
    sigma = SIGMA_TEMPLATE.format(
        title=title, rule_id=rule_id, description=description,
        refs=refs, tags=tags_lines, product=product or "unknown", level=level,
    )
    return spl, kql, sigma


@router.get("/clients/{client_id}/export/detection-rules.zip")
async def export_detection_rules_zip(client_id: str, user: dict = Depends(_require_auth)):
    from db import database as db
    items = await db.get_items(severity="CRITICAL,HIGH", client_id=client_id, limit=20)
    if not items:
        items = await db.get_items(severity="CRITICAL,HIGH", limit=20)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            slug = _slug(item.get("title", "unknown"))
            cves = item.get("cve_ids") or []
            if isinstance(cves, str):
                try:
                    cves = json.loads(cves)
                except Exception:
                    cves = []
            name = (cves[0] if cves else slug) or slug
            name = name.replace(":", "-")

            spl, kql, sigma = _make_rules(item)
            zf.writestr(f"splunk/{name}.spl", spl)
            zf.writestr(f"sentinel/{name}.kql", kql)
            zf.writestr(f"sigma/{name}.yml", sigma)

        # Add README
        zf.writestr("README.txt",
            "PhantomFeed Detection Rules\n"
            "============================\n"
            "splunk/   — Splunk SPL searches\n"
            "sentinel/ — Microsoft Sentinel KQL queries\n"
            "sigma/    — Sigma rules (convert with sigmac)\n"
        )

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="detection_rules.zip"'},
    )


# ---------------------------------------------------------------------------
# POST /clients/{id}/export/push-rules-github
# ---------------------------------------------------------------------------

@router.get("/clients/{client_id}/report.html")
async def client_report_html_preview(
    client_id: str,
    days: int = Query(30, ge=1, le=365),
    user: dict = Depends(_require_auth),
):
    """Browser-renderable HTML report with a Download PDF button at the top."""
    from db import database as db
    from reports.pdf_generator import generate_client_report_html, _get_report_extras
    from fastapi.responses import HTMLResponse
    client = await db.get_client(client_id)
    if not client:
        raise HTTPException(404, "Client not found")
    items = await db.get_items(client_id=client_id, limit=500, sort="risk")
    extras = await _get_report_extras(client_id, days)
    html = generate_client_report_html(client, items, days, extras=extras)
    btn = (f'<div style="padding:12px;background:#1a1a2e;text-align:center">'
           f'<a href="/api/v1/admin/clients/{client_id}/report.pdf?days={days}" '
           f'style="color:#fff;background:#6b46c1;padding:8px 20px;text-decoration:none;'
           f'font-family:sans-serif;font-size:13px">⬇ Download PDF</a></div>')
    html = html.replace("<body>", "<body>" + btn)
    return HTMLResponse(content=html)


@router.post("/clients/{client_id}/export/push-rules-github")
async def push_rules_github(client_id: str, body: dict, user: dict = Depends(_require_auth)):
    """Push detection rules to a GitHub repo via the Contents API."""
    repo = body.get("repo", "")
    token = body.get("token", "")
    branch = body.get("branch", "main")
    path = body.get("path", "detection-rules").strip("/")

    if not repo or not token:
        raise HTTPException(400, "repo and token are required")

    import httpx
    from db import database as db
    items = await db.get_items(severity="CRITICAL,HIGH", client_id=client_id, limit=20)
    if not items:
        items = await db.get_items(severity="CRITICAL,HIGH", limit=20)

    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
    }

    pushed = []
    errors = []

    async with httpx.AsyncClient(timeout=30) as client:
        for item in items:
            slug = _slug(item.get("title", "unknown"))
            cves = item.get("cve_ids") or []
            if isinstance(cves, str):
                try:
                    cves = json.loads(cves)
                except Exception:
                    cves = []
            name = (cves[0] if cves else slug) or slug
            name = name.replace(":", "-")

            spl, kql, sigma = _make_rules(item)
            files = [
                (f"{path}/splunk/{name}.spl", spl),
                (f"{path}/sentinel/{name}.kql", kql),
                (f"{path}/sigma/{name}.yml", sigma),
            ]
            for file_path, content in files:
                import base64
                encoded = base64.b64encode(content.encode()).decode()
                url = f"https://api.github.com/repos/{repo}/contents/{file_path}"
                # Check if file exists to get SHA for update
                sha = None
                r = await client.get(url, headers=headers, params={"ref": branch})
                if r.status_code == 200:
                    sha = r.json().get("sha")

                payload = {
                    "message": f"PhantomFeed: update {file_path}",
                    "content": encoded,
                    "branch": branch,
                }
                if sha:
                    payload["sha"] = sha

                r = await client.put(url, headers=headers, json=payload)
                if r.status_code in (200, 201):
                    pushed.append(file_path)
                else:
                    errors.append(f"{file_path}: {r.status_code} {r.text[:100]}")

    return {"pushed": len(pushed), "files": pushed, "errors": errors[:10]}
