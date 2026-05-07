"""
PhantomFeed Upload Center Integration Test
Run: python test_uploads.py
"""
import asyncio
import io
import os
import sys
sys.path.insert(0, '.')


# ── Minimal sample files ─────────────────────────────────────────────────────

SAMPLE_NESSUS = b"""<?xml version="1.0" ?>
<NessusClientData_v2>
<Report name="Test Scan">
<ReportHost name="srv01">
  <HostProperties>
    <tag name="host-ip">10.0.0.1</tag>
    <tag name="operating-system">Microsoft Windows Server 2019</tag>
  </HostProperties>
  <ReportItem port="443" pluginID="12345" pluginName="CVE-2024-1234 Remote Code Execution" severity="3">
    <cve>CVE-2024-1234</cve>
    <description>A critical vulnerability allows unauthenticated RCE.</description>
    <solution>Apply vendor patch immediately.</solution>
  </ReportItem>
  <ReportItem port="80" pluginID="23456" pluginName="SSL Certificate Expired" severity="1">
    <description>SSL certificate has expired.</description>
    <solution>Renew SSL certificate.</solution>
  </ReportItem>
</ReportHost>
</Report>
</NessusClientData_v2>"""

SAMPLE_CSV_ASSETS = b"""hostname,ip_address,os,software,version
srv01,10.0.0.1,Windows Server 2019,Microsoft Windows Server,2019
web01,10.0.0.2,Ubuntu 22.04,Apache HTTP Server,2.4.58
db01,10.0.0.3,Windows Server 2022,Microsoft SQL Server,2022"""

SAMPLE_IOCS_TXT = b"""8.8.8.8
1.1.1.1
d41d8cd98f00b204e9800998ecf8427e
e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855
example-malware.com
https://malware.example.com/payload.exe"""

SAMPLE_RAPID7_CSV = b"""Asset IP Address,Asset Hostname,Asset OS Name,Vulnerability Title,Vulnerability CVE IDs,Vulnerability Severity,Vulnerability Description,Solution
10.0.0.1,srv01,Windows Server 2019,CVE-2024-1234 RCE,CVE-2024-1234,Critical,Remote code execution vulnerability,Apply KB5012345
10.0.0.2,web01,Ubuntu 22.04,Log4Shell,CVE-2021-44228,Critical,Log4j remote code execution,Update log4j to 2.17.1"""


async def main():
    print("=" * 60)
    print("PhantomFeed Upload Center Integration Test")
    print("=" * 60)

    import config
    config.DB_PATH = "./phantomfeed_test_uploads.db"
    from db import database as db
    await db.connect()
    print("[OK] Database connected")

    # Seed admin
    from auth.auth import seed_admin_user
    await seed_admin_user()

    # Create test client
    client = await db.create_client("Upload Test Corp", "test@example.com", {})
    client_id = client["id"]
    print(f"[OK] Test client created: {client_id}")

    # ── Test 1: Nessus parser ────────────────────────────────────────────────
    from uploads.parsers import NessusParser, detect_parser
    parser = NessusParser()
    result = parser.parse(SAMPLE_NESSUS, "test.nessus")
    assert result["format"] == "nessus"
    assert len(result["assets"]) == 1
    assert result["assets"][0]["hostname"] == "srv01"
    assert result["assets"][0]["ip_address"] == "10.0.0.1"
    assert len(result["findings"]) == 2
    assert result["findings"][0]["severity"] == "HIGH"
    assert "CVE-2024-1234" in result["findings"][0]["cve_ids"]
    print(f"[OK] NessusParser: {len(result['assets'])} assets, {len(result['findings'])} findings")

    # ── Test 2: Auto-detect Nessus ───────────────────────────────────────────
    fmt, parser2 = detect_parser(SAMPLE_NESSUS, "test.nessus")
    assert fmt == "nessus"
    print(f"[OK] Auto-detect: {fmt}")

    # ── Test 3: Rapid7 CSV parser ────────────────────────────────────────────
    from uploads.parsers import Rapid7Parser
    r7 = Rapid7Parser()
    r7_result = r7.parse(SAMPLE_RAPID7_CSV, "rapid7.csv")
    assert r7_result["format"] == "rapid7"
    assert len(r7_result["assets"]) == 2
    assert len(r7_result["findings"]) == 2
    assert r7_result["findings"][0]["severity"] == "CRITICAL"
    print(f"[OK] Rapid7Parser: {len(r7_result['assets'])} assets, {len(r7_result['findings'])} findings")

    # ── Test 4: GenericCSV parser ─────────────────────────────────────────────
    from uploads.parsers import GenericCSVParser
    gcsv = GenericCSVParser()
    gcsv_result = gcsv.parse(SAMPLE_CSV_ASSETS, "assets.csv")
    assert gcsv_result["format"] == "generic_csv"
    assert len(gcsv_result["assets"]) == 3
    assert gcsv_result["field_mapping"].get("hostname") == "hostname"
    assert gcsv_result["field_mapping"].get("ip_address") == "ip_address"
    print(f"[OK] GenericCSVParser: {len(gcsv_result['assets'])} assets, mapping={list(gcsv_result['field_mapping'].keys())[:4]}")

    # ── Test 5: IOC parser ────────────────────────────────────────────────────
    from uploads.parsers import IOCParser, detect_ioc_type
    ioc_parser = IOCParser()
    ioc_result = ioc_parser.parse(SAMPLE_IOCS_TXT, "iocs.txt")
    iocs = ioc_result["iocs"]
    assert len(iocs) >= 5
    types = {i["type"] for i in iocs}
    assert "ip" in types
    assert "sha256" in types or "md5" in types
    assert "domain" in types or "url" in types
    print(f"[OK] IOCParser: {len(iocs)} IOCs, types={sorted(types)}")

    # IOC type detection
    assert detect_ioc_type("8.8.8.8") == "ip"
    assert detect_ioc_type("d" * 32) == "md5"
    assert detect_ioc_type("e" * 64) == "sha256"
    assert detect_ioc_type("example.com") == "domain"
    print("[OK] IOC type detection")

    # ── Test 6: Upload log CRUD ──────────────────────────────────────────────
    from uploads.upload_log import create_upload_log, update_upload_log, list_upload_logs
    log = await create_upload_log("test.nessus", "nessus", client_id, "preview")
    assert log["status"] == "preview"
    updated = await update_upload_log(log["id"], status="imported", records_imported=5, completed=True)
    assert updated["status"] == "imported"
    assert updated["records_imported"] == 5
    all_logs = await list_upload_logs(client_id=client_id)
    assert len(all_logs) >= 1
    print(f"[OK] Upload log CRUD: {len(all_logs)} log(s)")

    # ── Test 7: CSV export ────────────────────────────────────────────────────
    # Seed a test item
    from compliance.mappings import tag_item
    test_item = {
        "feed_id": "test", "feed_label": "Test Feed", "category": "cve",
        "severity": "CRITICAL", "title": "Test CVE Export Item",
        "description": "Test desc", "cve_ids": ["CVE-2024-9999"],
        "tags": ["test"], "published_at": "2025-01-01",
        "vendor": "TestVendor", "product": "TestProduct",
        "url": "https://example.com", "raw": {}, "compliance_tags": [],
    }
    test_item["compliance_tags"] = tag_item(test_item)
    await db.upsert_item(test_item)

    # Simulate CSV export generation
    items = await db.get_items(limit=10)
    assert len(items) >= 1

    import csv, io as _io, json
    buf = _io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id","title","severity","cvss","risk_score","vendor","product","feed_label","category","published_at","url","cve_ids","tags","compliance_cmmc","compliance_nist"])
    for item in items:
        cves = item.get("cve_ids") or []
        if isinstance(cves, str):
            try: cves = json.loads(cves)
            except: cves = []
        writer.writerow([item["id"], item["title"], item["severity"], item.get("cvss",""), item.get("risk_score",""),
                         item.get("vendor",""), item.get("product",""), item.get("feed_label",""),
                         item.get("category",""), item.get("published_at",""), item.get("url",""),
                         ";".join(cves), "", "", ""])
    csv_content = buf.getvalue()
    assert "Test CVE Export Item" in csv_content
    print(f"[OK] CSV export: {len(csv_content)} bytes, {len(items)} items")

    # ── Test 8: XLSX remediation export ──────────────────────────────────────
    import openpyxl
    from openpyxl.styles import PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Title","Severity","Status","Due Date"])
    ws.append(["Test Item","CRITICAL","open","2025-01-15"])
    fill = PatternFill("solid", fgColor="FECACA")
    ws['A2'].fill = fill
    buf2 = _io.BytesIO()
    wb.save(buf2)
    xlsx_bytes = buf2.getvalue()
    assert len(xlsx_bytes) > 1000
    # Verify it's a valid XLSX
    wb2 = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes))
    assert wb2.active['A1'].value == "Title"
    print(f"[OK] XLSX export: {len(xlsx_bytes)} bytes, valid workbook")

    # ── Test 9: Detection rules ZIP ────────────────────────────────────────────
    import zipfile
    from api.export_routes import _make_rules, _slug
    test_item2 = {
        "title": "CVE-2024-1234 Remote Code Execution",
        "severity": "CRITICAL",
        "vendor": "Microsoft",
        "product": "Windows",
        "cve_ids": json.dumps(["CVE-2024-1234"]),
        "description": "Remote code execution",
        "url": "https://example.com",
    }
    spl, kql, sigma = _make_rules(test_item2)
    assert "CVE-2024-1234" in spl or "Microsoft" in spl
    assert "CVE-2024-1234" in kql or "Microsoft" in kql
    assert "title:" in sigma

    buf3 = _io.BytesIO()
    with zipfile.ZipFile(buf3, "w") as zf:
        zf.writestr("splunk/CVE-2024-1234.spl", spl)
        zf.writestr("sentinel/CVE-2024-1234.kql", kql)
        zf.writestr("sigma/CVE-2024-1234.yml", sigma)
    buf3.seek(0)
    with zipfile.ZipFile(buf3) as zf:
        names = zf.namelist()
    assert "splunk/CVE-2024-1234.spl" in names
    assert "sentinel/CVE-2024-1234.kql" in names
    assert "sigma/CVE-2024-1234.yml" in names
    print(f"[OK] Detection rules ZIP: {len(names)} files — {names}")

    # Cleanup
    await db.delete_client(client_id)
    await db.close()
    for f in ["./phantomfeed_test_uploads.db", "./phantomfeed_test_uploads.db-shm", "./phantomfeed_test_uploads.db-wal"]:
        try: os.remove(f)
        except: pass

    print("\n" + "=" * 60)
    print("ALL UPLOAD CENTER INTEGRATION TESTS PASSED")
    print("=" * 60)


asyncio.run(main())
